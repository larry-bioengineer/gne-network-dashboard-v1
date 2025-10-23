from flask import Blueprint, jsonify, request, Response
from api.models import ResponseModel
import subprocess
import pandas as pd
import os
import time
import threading
import json
import socket
from dotenv import load_dotenv
from service.SSHConnection import reset_port_poe, retrieve_ssh_info_from_config

from api.util import log

# Create a Blueprint for API routes
api_bp = Blueprint('api', __name__, url_prefix='/api')

def get_locations_from_data_file():
    """
    Get all locations from the data.xlsx file.
    This is a shared function to avoid duplication between endpoints.
    
    Returns:
        tuple: (success: bool, result: dict or str)
               - If success=True, result contains {'locations': list, 'df_clean': DataFrame}
               - If success=False, result contains error message
    """
    try:
        # Get the project root directory (parent of api directory)
        project_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        excel_path = os.path.join(project_root, 'config_file', 'data.xlsx')
        
        df = pd.read_excel(excel_path, sheet_name='Hardware list')
        
        # Return error if Location or IP is not found
        if 'Location' not in df.columns or 'IP' not in df.columns:
            return False, "Location or IP column not found in data file"
        
        # Filter out NaN values
        df_clean = df.dropna(subset=['Location', 'IP'])
        locations = df_clean['Location'].unique().tolist()
        
        if not locations:
            return False, "No locations found in data file"
        
        return True, {'locations': locations, 'df_clean': df_clean}
        
    except Exception as e:
        return False, f"Failed to read data file: {str(e)}"

def analyze_ssh_error(ip, port, timeout):
    """
    Analyze potential SSH connection issues and provide diagnostic information
    """
    diagnostics = {
        "connection_test": "pending",
        "ssh_port_open": "pending", 
        "hostname_resolution": "pending",
        "recommendations": []
    }
    
    try:
        # Test hostname resolution
        socket.gethostbyname(ip)
        diagnostics["hostname_resolution"] = "success"
    except socket.gaierror:
        diagnostics["hostname_resolution"] = "failed"
        diagnostics["recommendations"].append("Check if the IP address or hostname is correct")
    
    try:
        # Test if SSH port is open
        sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
        sock.settimeout(5)
        result = sock.connect_ex((ip, port))
        sock.close()
        
        if result == 0:
            diagnostics["ssh_port_open"] = "open"
            diagnostics["connection_test"] = "success"
        else:
            diagnostics["ssh_port_open"] = "closed"
            diagnostics["connection_test"] = "failed"
            diagnostics["recommendations"].append(f"SSH port {port} appears to be closed or filtered")
    except Exception as e:
        diagnostics["connection_test"] = "error"
        diagnostics["recommendations"].append(f"Unable to test connection: {str(e)}")
    
    return diagnostics

# Admin endpoints
@api_bp.route('/admin/health', methods=['GET'])
def health_check():
    """Health check endpoint that returns success status"""
    response = ResponseModel(
        success=True,
        message="API is healthy and running",
        data={"status": "ok", "service": "SSH Automation Server"}
    )
    return jsonify(response.to_dict())

@api_bp.route('/admin/status', methods=['GET'])
def get_status():
    """Status endpoint that returns current system status"""
    response = ResponseModel(
        success=True,
        message="System status retrieved successfully",
        data={
            "system": "SSH Automation Server",
            "version": "1.0.0",
            "uptime": "running"
        }
    )
    return jsonify(response.to_dict())

# Network endpoints
@api_bp.route('/network/ping', methods=['POST'])
def ping():
    """
    Ping a given network interface

    Args:
        interface: The network interface to ping

    Returns:
        A JSON response with the ping result
    """
    data = request.json
    interface = data.get('interface')

    if not interface:
        return jsonify({'error': 'Interface is required'}), 400

    # ping the interface
    # Set reasonable default ping count and timeout (3 packets, 3s per packet)
    ping_count = 3
    ping_timeout = 3  # seconds per packet
    ping_result = subprocess.run(
        ['ping', interface], 
        capture_output=True, 
        text=True
    )
    
    return jsonify({'result': ping_result.stdout})

@api_bp.route('/network/get_ip_and_location', methods=['GET'])
def get_ip_and_location():
    """
    Get the IP and location from the data.xlsx file
    """
    success, result = get_locations_from_data_file()
    
    if not success:
        return jsonify({'error': result}), 500
    
    # Extract the clean dataframe from the result
    df_clean = result['df_clean']
    
    return jsonify({
        'Location': df_clean['Location'].values.tolist(), 
        'IP': df_clean['IP'].values.tolist()
    })

@api_bp.route('/network/ping_specific_location', methods=['POST'])
def ping_specific_location():
    """
    Ping a given network interface to a specific location

    Args:
        interface: The network interface to ping
        location: The location to ping

    Returns:
        A JSON response with the ping result
    """
    data = request.json
    
    try:
        # open the data.xlsx file
        df = pd.read_excel('config_file/data.xlsx', sheet_name='Hardware List')
        
        # Filter out NaN values to prevent issues
        df_clean = df.dropna(subset=['Location', 'IP'])

        # get the interface and location
        interface = data.get('interface')
        location = data.get('location')

        if not interface or not location:
            return jsonify({'error': 'Both interface and location are required'}), 400

        # Find the IP for the specific location
        location_data = df_clean[df_clean['Location'] == location]
        if location_data.empty:
            return jsonify({'error': f'Location "{location}" not found'}), 404

        target_ip = location_data['IP'].iloc[0]
        
        # Ping the specific IP
        ping_result = subprocess.run(['ping', target_ip], capture_output=True, text=True)
        
        return jsonify({
            'result': ping_result.stdout,
            'location': location,
            'target_ip': target_ip
        })
        
    except Exception as e:
        return jsonify({'error': f'Failed to ping location: {str(e)}'}), 500

@api_bp.route('/network/ping_sse', methods=['POST'])
def ping_sse():
    """
    Ping a given network interface using Server-Sent Events (SSE)
    
    Args:
        interface: The network interface to ping
        count: Number of ping packets to send (optional, defaults to 4)
        
    Returns:
        A Server-Sent Events stream with ping results
    """
    data = request.json
    interface = data.get('interface')
    count = data.get('count', 4)
    
    if not interface:
        return jsonify({'error': 'Interface is required'}), 400
    
    def generate_ping_events():
        try:
            # Send initial event
            message_text = f'Starting ping to {interface}'
            yield f"data: {json.dumps({'type': 'start', 'message': message_text, 'timestamp': time.time()})}\n\n"
            
            # Start ping process
            ping_process = subprocess.Popen(
                ['ping', interface],
                stdout=subprocess.PIPE,
                stderr=subprocess.PIPE,
                text=True,
                bufsize=1,
                universal_newlines=True
            )
            
            # Read ping output line by line
            for line in iter(ping_process.stdout.readline, ''):
                if line:
                    # Send each ping line as an event
                    yield f"data: {json.dumps({'type': 'ping_line', 'data': line.strip(), 'timestamp': time.time()})}\n\n"
                    time.sleep(0.1)  # Small delay to prevent overwhelming the client
            
            # Wait for process to complete
            ping_process.wait()
            
            # Send completion event
            if ping_process.returncode == 0:
                yield f"data: {json.dumps({'type': 'complete', 'message': 'Ping completed successfully', 'timestamp': time.time()})}\n\n"
            else:
                yield f"data: {json.dumps({'type': 'error', 'message': 'Ping failed', 'timestamp': time.time()})}\n\n"
                
        except Exception as e:
            error_message = f'Error during ping: {str(e)}'
            yield f"data: {json.dumps({'type': 'error', 'message': error_message, 'timestamp': time.time()})}\n\n"
    
    return Response(
        generate_ping_events(),
        mimetype='text/event-stream',
        headers={
            'Cache-Control': 'no-cache',
            'Connection': 'keep-alive',
            'Access-Control-Allow-Origin': '*',
            'Access-Control-Allow-Headers': 'Cache-Control'
        }
    )

@api_bp.route('/network/ping_sse_location', methods=['POST'])
def ping_sse_location():
    """
    Ping a specific location using Server-Sent Events (SSE)
    
    Args:
        location: The location to ping
        count: Number of ping packets to send (optional, defaults to 4)
        
    Returns:
        A Server-Sent Events stream with ping results
    """
    data = request.json
    location = data.get('location')
    count = data.get('count', 4)
    
    if not location:
        return jsonify({'error': 'Location is required'}), 400
    
    def generate_ping_events():
        try:
            # Load location data
            df = pd.read_excel('config_file/data.xlsx', sheet_name='Hardware list')
            df_clean = df.dropna(subset=['Location', 'IP'])
            
            # Find the IP for the specific location
            location_data = df_clean[df_clean['Location'] == location]
            if location_data.empty:
                error_message = f'Location "{location}" not found'
                yield f"data: {json.dumps({'type': 'error', 'message': error_message, 'timestamp': time.time()})}\n\n"
                return
            
            target_ip = location_data['IP'].iloc[0]
            
            # Send initial event
            message_text = f'Starting ping to {location} ({target_ip})'
            yield f"data: {json.dumps({'type': 'start', 'message': message_text, 'location': location, 'target_ip': target_ip, 'timestamp': time.time()})}\n\n"
            
            # Start ping process
            ping_process = subprocess.Popen(
                ['ping', target_ip],
                stdout=subprocess.PIPE,
                stderr=subprocess.PIPE,
                text=True,
                bufsize=1,
                universal_newlines=True
            )
            
            # Read ping output line by line
            for line in iter(ping_process.stdout.readline, ''):
                if line:
                    # Send each ping line as an event
                    yield f"data: {json.dumps({'type': 'ping_line', 'data': line.strip(), 'location': location, 'target_ip': target_ip, 'timestamp': time.time()})}\n\n"
                    time.sleep(0.1)  # Small delay to prevent overwhelming the client
            
            # Wait for process to complete
            ping_process.wait()
            
            # Send completion event
            if ping_process.returncode == 0:
                success_message = f'Ping to {location} completed successfully'
                yield f"data: {json.dumps({'type': 'complete', 'message': success_message, 'location': location, 'target_ip': target_ip, 'timestamp': time.time()})}\n\n"
            else:
                error_message = f'Ping to {location} failed'
                yield f"data: {json.dumps({'type': 'error', 'message': error_message, 'location': location, 'target_ip': target_ip, 'timestamp': time.time()})}\n\n"
                
        except Exception as e:
            error_message = f'Error during ping: {str(e)}'
            yield f"data: {json.dumps({'type': 'error', 'message': error_message, 'timestamp': time.time()})}\n\n"
    
    return Response(
        generate_ping_events(),
        mimetype='text/event-stream',
        headers={
            'Cache-Control': 'no-cache',
            'Connection': 'keep-alive',
            'Access-Control-Allow-Origin': '*',
            'Access-Control-Allow-Headers': 'Cache-Control'
        }
    )

def verify_port_connectivity(ip_address, timeout=3):
    """
    Verify that a port is back up after reset by pinging the IP address.
    
    Args:
        ip_address: The IP address to ping
        timeout: Timeout in seconds for the ping operation
        
    Returns:
        bool: True if ping is successful, False otherwise
    """
    try:
        # Force English output to avoid localization issues
        env = os.environ.copy()
        env['LC_ALL'] = 'C'
        
        ping_result = subprocess.run(
            ['ping', ip_address], 
            capture_output=True, 
            text=True,
            timeout=timeout + 5,
            env=env
        )
        
        # Check for universal success indicators that work across languages
        ping_success = False
        if ping_result.returncode == 0:
            ping_output = ping_result.stdout.lower()
            # Look for TTL (Time To Live) which indicates successful ping response
            if 'ttl=' in ping_output:
                ping_success = True
            # Alternative: look for timing indicators (ms) which also indicate success
            elif 'ms' in ping_output and ('time=' in ping_output or '時間=' in ping_output):
                ping_success = True
        
        return ping_success
        
    except subprocess.TimeoutExpired:
        return False
    except Exception as e:
        print(f"Error during port verification ping: {e}")
        return False

@api_bp.route('/network/reset_port', methods=['POST'])
def reset_port():
    """
    Reset the port of a given location name by disabling and enabling PoE
    
    Args:
        request.json (dict): JSON payload containing:
            - locName (str, required): The location name to reset the port for
            - port (int, optional): The port number to reset (e.g., 16 for ge-0/0/16). 
                                   Defaults to 22 if not provided.
            - timeout (int, optional): Connection timeout in seconds. Defaults to 10.
    
    Returns:
        ResponseModel: JSON response containing:
            - success (bool): Whether the operation was successful
            - message (str): Human-readable message describing the result
            - data (dict, optional): Additional data including:
                - locName (str): The location name
                - port (int): The port number that was reset
                - status (str): Status of the reset operation
    
    Raises:
        400: If location name is missing from request
        500: If SSH credentials are not configured or invalid
        500: If port reset operation fails
        500: If any other error occurs during execution
    
    Environment Variables Required:
        - SSH_USERNAME: SSH username for switch authentication
        - SSH_PASSWORD: SSH password for switch authentication
        - SSH_PORT: SSH port number (optional, defaults to 22)
    
    """
    try:
        # Load environment variables
        load_dotenv()
        
        data = request.json
        locName = data.get('locName')
        timeout = data.get('timeout', 10)  # Default timeout of 10 seconds

        if not locName:
            response = ResponseModel(
                success=False,
                message="Location name is required",
                data=None
            )
            return jsonify(response.to_dict()), 400

        # INSERT PORT NUMBER LOOKUP CODE HERE
        config = retrieve_ssh_info_from_config(locName)
        
        # Validate timeout parameter
        try:
            timeout = int(timeout)
            if timeout <= 0 or timeout > 300:  # Allow 1-300 seconds
                raise ValueError("Timeout must be between 1 and 300 seconds")
        except (ValueError, TypeError):
            response = ResponseModel(
                success=False,
                message="Invalid timeout value. Must be an integer between 1 and 300 seconds",
                data=None
            )
            return jsonify(response.to_dict()), 400
        
        if not config:
            response = ResponseModel(
                success=False,
                message="SSH Configuration not found",
                data=None
            )
            return jsonify(response.to_dict()), 400
        
        # Get SSH credentials from environment variables
        ssh_username = os.getenv('SSH_USERNAME')
        ssh_password = os.getenv('SSH_PASSWORD')
        ssh_port = config['port']
        
        # Validate that all required credentials are provided
        if not ssh_username or not ssh_password:
            response = ResponseModel(
                success=False,
                message="SSH credentials not configured. Please set SSH_USERNAME and SSH_PASSWORD in .env file",
                data=None
            )
            return jsonify(response.to_dict()), 500
        
        
        try:
            ssh_port = int(ssh_port)
        except ValueError:
            response = ResponseModel(
                success=False,
                message="Invalid SSH_PORT in config file. Must be a valid integer",
                data=None
            )
            return jsonify(response.to_dict()), 500
        
        # Store the target_port before overwriting config
        target_port = config['target_port']
        
        # Create SSH configuration
        ssh_config = {
            'hostname': config['hostname'],
            'username': ssh_username,
            'password': ssh_password,
            'port': ssh_port
        }
        
        # Execute port reset with timeout
        success = reset_port_poe(ssh_config, target_port, timeout)
        
        if success:
            response = ResponseModel(   
                success=True,
                message=f"Port ge-0/0/{target_port} reset successfully for {locName}",
                data={
                    "locName": locName,
                    "port": target_port,
                    "status": f"reset_completed for {locName}"
                }
            )
            return jsonify(response.to_dict())
        else:
            # Perform SSH connection diagnostics
            diagnostics = analyze_ssh_error(ssh_config['hostname'], ssh_port, timeout)
            
            # Determine error type based on diagnostics
            if diagnostics["hostname_resolution"] == "failed":
                error_type = "DNS_RESOLUTION_FAILED"
                error_message = f"🔍 Cannot resolve hostname/IP: {ssh_config['hostname']}. Please verify the IP address is correct."
            elif diagnostics["ssh_port_open"] == "closed":
                error_type = "SSH_PORT_CLOSED"
                error_message = f"🚫 SSH connection failed: Port {ssh_port} is closed or filtered. SSH service may not be running."
            elif diagnostics["connection_test"] == "success":
                error_type = "SSH_AUTHENTICATION_FAILED"
                error_message = f"🔐 SSH connection established but authentication failed. Please check credentials."
            else:
                error_type = "SSH_CONNECTION_TIMEOUT"
                error_message = f"⏱️ SSH connection timeout after {timeout} seconds. Check network connectivity and firewall settings."
            
            response = ResponseModel(
                success=False,
                message=error_message,
                data={
                    "locName": locName,
                    "port": target_port,
                    "ssh_port": ssh_port,
                    "timeout": timeout,
                    "status": "reset_failed",
                    "error_type": error_type,
                    "diagnostics": diagnostics
                }
            )
            return jsonify(response.to_dict()), 500
            
    except Exception as e:
        # Handle the case where config might not be initialized
        diagnostics = None
        port_number = None
        hostname = None
        
        try:
            if 'ssh_config' in locals() and ssh_config:
                hostname = ssh_config['hostname']
                ssh_port = ssh_config['port'] if 'port' in ssh_config else 22
                diagnostics = analyze_ssh_error(hostname, int(ssh_port) if ssh_port else 22, timeout)
                port_number = target_port if 'target_port' in locals() else None
            elif 'config' in locals() and config:
                hostname = config['hostname']
                port_number = config['target_port']
                ssh_port = config['port'] if 'port' in config else 22
                diagnostics = analyze_ssh_error(hostname, int(ssh_port) if ssh_port else 22, timeout)
            else:
                # Use the retrieved config data for diagnostics
                retrieved_config = retrieve_ssh_info_from_config(locName)
                if retrieved_config:
                    hostname = retrieved_config['hostname']
                    port_number = retrieved_config['target_port']
                    diagnostics = analyze_ssh_error(hostname, retrieved_config['port'], timeout)
        except:
            # If we can't get diagnostics, just proceed without them
            pass
        
        response = ResponseModel(
            success=False,
            message=f"❌ Unexpected error during port reset: {str(e)}",
            data={
                "locName": locName,
                "port": port_number,
                "status": "error",
                "error_type": "UNEXPECTED_ERROR",
                "diagnostics": diagnostics
            }
        )
        return jsonify(response.to_dict()), 500


@api_bp.route('/network/reset_all_locations', methods=['POST'])
def reset_all_locations():
    """
    Reset all locations from data.xlsx file
    
    Args:
        request.json (dict, optional): JSON payload containing:
            - timeout (int, optional): Connection timeout in seconds for each reset
                Defaults to 10
    
    Returns:
        ResponseModel: JSON response containing:
            - success (bool): Whether the operation was successful for all locations
            - message (str): Human-readable summary message
            - data (dict): Details of the operation including:
                - successful_locations: List of locations successfully reset
                - failed_locations: List of locations that failed
                - total: Total number of locations processed
                - success_count: Number of successful resets
                - failed_count: Number of failed resets
    """
    try:
        # Load environment variables first
        load_dotenv()
        
        data = request.json if request.json else {}
        timeout = data.get('timeout', 10)  # Default timeout of 10 seconds
        
        # Validate timeout parameter
        try:
            timeout = int(timeout)
            if timeout <= 0 or timeout > 300:  # Allow 1-300 seconds
                raise ValueError("Timeout must be between 1 and 300 seconds")
        except (ValueError, TypeError):
            response = ResponseModel(
                success=False,
                message="Invalid timeout value. Must be an integer between 1 and 300 seconds",
                data=None
            )
            return jsonify(response.to_dict()), 400
        
        # Check if SSH credentials are available
        ssh_username = os.getenv('SSH_USERNAME')
        ssh_password = os.getenv('SSH_PASSWORD')
        
        if not ssh_username or not ssh_password:
            response = ResponseModel(
                success=False,
                message="SSH credentials not configured. Please set SSH_USERNAME and SSH_PASSWORD in .env file",
                data=None
            )
            return jsonify(response.to_dict()), 500
        
        # Get all locations from data.xlsx using shared function
        success, result = get_locations_from_data_file()
        
        if not success:
            response = ResponseModel(
                success=False,
                message=result,
                data=None
            )
            return jsonify(response.to_dict()), 500
        
        # Extract locations from the result
        locations = result['locations']
        
        # Initialize progress tracking
        successful_locations = []
        failed_locations = []
        total_processed = len(locations)
        
        # Reset each location
        for i, location in enumerate(locations):
            try:
                # Use same logic as network/reset_port for each location
                # Get SSH configuration for this location
                config = retrieve_ssh_info_from_config(location)
                
                if not config:
                    failed_locations.append({
                        'location': location,
                        'error': 'SSH Configuration not found'
                    })
                    continue
                
                # Create SSH configuration (same as reset_port)
                ssh_config = {
                    'hostname': config['hostname'],
                    'username': ssh_username,
                    'password': ssh_password,
                    'port': int(config['port'])
                }
                
                target_port = config['target_port']
                
                # Execute port reset (same as reset_port)
                success = reset_port_poe(ssh_config, target_port, timeout)
                
                if success:
                    successful_locations.append(location)
                else:
                    # Perform diagnostics (same as reset_port)
                    diagnostics = analyze_ssh_error(ssh_config['hostname'], config['port'], timeout)
                    failed_locations.append({
                        'location': location,
                        'error': 'SSH connection or reset failed',
                        'diagnostics': diagnostics
                    })
                
                # Small delay between resets to avoid overwhelming switches
                if i < len(locations) - 1:
                    time.sleep(2)
                    
            except Exception as e:
                failed_locations.append({
                    'location': location,
                    'error': str(e)
                })
        
        success_count = len(successful_locations)
        failed_count = len(failed_locations)
        
        # Build response
        if failed_count == 0:
            overall_success = True
            summary_message = f"Successfully reset {success_count} out of {total_processed} locations"
        else:
            overall_success = len(successful_locations) > 0  # True if at least one succeeded
            summary_message = f"Reset operation completed: {success_count} successful, {failed_count} failed out of {total_processed} locations"
        
        response = ResponseModel(
            success=overall_success,
            message=summary_message,
            data={
                'successful_locations': successful_locations,
                'failed_locations': failed_locations,
                'total': total_processed,
                'success_count': success_count,
                'failed_count': failed_count
            }
        )
        
        # Return 200 if any success, 500 if all failed
        status_code = 200 if overall_success else 500
        return jsonify(response.to_dict()), status_code
        
    except Exception as e:
        response = ResponseModel(
            success=False,
            message=f"Unexpected error during reset all operation: {str(e)}",
            data={'error_type': 'UNEXPECTED_ERROR'}
        )
        return jsonify(response.to_dict()), 500


@api_bp.route('/network/reset_down_port_only', methods=['POST'])
def reset_down_port_only():
    """
    Go through the data.xlsx. Ping the IP address. If the IP address is not reachable, 
    reset the port of a given location name by disabling PoE
    
    Args:
        request.json (dict, optional): JSON payload containing:
            - timeout (int, optional): Connection timeout in seconds for each ping and reset
                Defaults to 10
    
    Returns:
        ResponseModel: JSON response containing:
            - success (bool): Whether the operation was successful
            - message (str): Human-readable message describing the result
            - data (dict): Details of the operation including:
                - total_checked: Total number of locations checked
                - down_locations: List of locations that were unreachable (ping failed)
                - reset_attempted: Locations where port reset was attempted
                - reset_successful: Locations successfully reset
                - reset_failed: Locations where reset failed
                - not_down: Locations that were reachable (no action needed)
    
    """
    try:
        # Load environment variables
        load_dotenv()
        
        data = request.json if request.json else {}
        timeout = data.get('timeout', 10)  # Default timeout of 10 seconds
        
        # Validate timeout parameter
        try:
            timeout = int(timeout)
            if timeout <= 0 or timeout > 300:  # Allow 1-300 seconds
                raise ValueError("Timeout must be between 1 and 300 seconds")
        except (ValueError, TypeError):
            response = ResponseModel(
                success=False,
                message="Invalid timeout value. Must be an integer between 1 and 300 seconds",
                data=None
            )
            return jsonify(response.to_dict()), 400
        
        # Check if SSH credentials are available
        ssh_username = os.getenv('SSH_USERNAME')
        ssh_password = os.getenv('SSH_PASSWORD')
        
        if not ssh_username or not ssh_password:
            response = ResponseModel(
                success=False,
                message="SSH credentials not configured. Please set SSH_USERNAME and SSH_PASSWORD in .env file",
                data=None
            )
            return jsonify(response.to_dict()), 500
        
        # Get all locations from data.xlsx using the shared function
        success, result = get_locations_from_data_file()
        
        if not success:
            response = ResponseModel(
                success=False,
                message=result,
                data=None
            )
            return jsonify(response.to_dict()), 500
        
        # Extract locations and dataframe from the result
        locations = result['locations']
        df_clean = result['df_clean']
        
        # Process all locations
        total_checked = len(locations)
        down_locations = []
        reset_attempted = []
        reset_successful = []
        reset_failed = []
        not_down = []
        
        # Check each location
        for location in locations:
            try:
                # Get the IP for this location
                location_data = df_clean[df_clean['Location'] == location]
                if location_data.empty:
                    continue
                
                ip_address = location_data['IP'].iloc[0]
                
                # Convert to string in case it's not already
                ip_address = str(ip_address).strip()
                
                # Ping the IP address with more robust parameters
                print(f"Pinging {location} at {ip_address} with 3 packets, 3s timeout...")
                # Force English output to avoid localization issues
                env = os.environ.copy()
                env['LC_ALL'] = 'C'
                ping_result = subprocess.run(
                    ['ping', ip_address], 
                    capture_output=True, 
                    text=True,
                    timeout=timeout + 5,  # Add buffer to ping timeout
                    env=env
                )
                
                # More robust ping result validation using universal indicators
                ping_success = False
                if ping_result.returncode == 0:
                    # Check for universal success indicators that work across languages
                    ping_output = ping_result.stdout.lower()
                    # Look for TTL (Time To Live) which indicates successful ping response
                    if 'ttl=' in ping_output:
                        ping_success = True
                    # Alternative: look for timing indicators (ms) which also indicate success
                    elif 'ms' in ping_output and ('time=' in ping_output or '時間=' in ping_output):
                        ping_success = True
                
                # Check if ping was successful
                if ping_success:
                    # IP is reachable - no action needed
                    not_down.append({
                        'location': location, 
                        'ip': ip_address,
                        'status': 'reachable'
                    })
                else:
                    # IP is not reachable - attempt port reset
                    down_locations.append({
                        'location': location, 
                        'ip': ip_address,
                        'status': 'unreachable'
                    })
                    
                    # Log ping failure details for debugging
                    print(f"Ping failed for {location} - Return code: {ping_result.returncode}")
                    print(f"Ping output: {ping_result.stdout[:200]}...")
                    
                    try:
                        # Get SSH configuration for this location
                        config = retrieve_ssh_info_from_config(location)
                        
                        if not config:
                            reset_failed.append({
                                'location': location,
                                'error': 'SSH configuration not found'
                            })
                        else:
                            # Create SSH configuration
                            ssh_config = {
                                'hostname': config['hostname'],
                                'username': ssh_username,
                                'password': ssh_password,
                                'port': int(config['port'])
                            }
                            
                            target_port = config['target_port']
                            reset_attempted.append({
                                'location': location,
                                'ip': ip_address,
                                'target_port': target_port
                            })
                            
                            # Execute port reset
                            success = reset_port_poe(ssh_config, target_port, timeout)
                            
                            if success:
                                reset_successful.append({
                                    'location': location,
                                    'ip': ip_address,
                                    'target_port': target_port,
                                    'status': 'reset_successful',
                                    'verified': False
                                })
                                print(f"Port {target_port} reset completed for {location} - verification will be performed after all resets")
                            else:
                                reset_failed.append({
                                    'location': location,
                                    'ip': ip_address,
                                    'target_port': target_port,
                                    'error': 'Port reset failed'
                                })
                    
                    except Exception as e:
                        reset_failed.append({
                            'location': location,
                            'error': f'Exception during reset: {str(e)}'
                        })
                        
            except subprocess.TimeoutExpired:
                # Ping timed out - treat as unreachable
                down_locations.append({
                    'location': location, 
                    'ip': ip_address,
                    'status': 'ping_timeout'
                })
                
            except Exception as e:
                # Error during processing
                reset_failed.append({
                    'location': location,
                    'error': f'Error during processing: {str(e)}'
                })
        
        # Perform batch verification after all resets are completed
        successful_resets = len(reset_successful)
        total_down = len(down_locations)
        
        if successful_resets > 0:
            # Wait for devices to fully resume after resets
            batch_verification_delay = int(os.getenv('BATCH_VERIFICATION_DELAY_SECONDS', '30'))
            print(f"All {successful_resets} port resets completed. Waiting {batch_verification_delay} seconds for devices to resume before verification...")
            time.sleep(batch_verification_delay)
            
            # Now verify all reset ports
            verified_count = 0
            for reset_entry in reset_successful:
                location = reset_entry['location']
                ip_address = reset_entry['ip']
                target_port = reset_entry['target_port']
                
                print(f"Verifying {location} ({ip_address})...")
                verification_success = verify_port_connectivity(ip_address, timeout)
                
                if verification_success:
                    reset_entry['verified'] = True
                    verified_count += 1
                    print(f"{location} is now reachable")
                else:
                    reset_entry['verified'] = False
                    print(f"{location} is still unreachable")
        
        # Build response message
        successful_resets = len(reset_successful)
        total_down = len(down_locations)
        
        # Count verified vs unverified resets
        verified_resets = len([r for r in reset_successful if r.get('verified', False)])
        unverified_resets = len([r for r in reset_successful if not r.get('verified', False)])
        
        if successful_resets > 0:
            if verified_resets == successful_resets:
                message = f"Completed port reset scan: {successful_resets} ports reset and all verified successfully out of {total_down} unreachable locations"
            elif verified_resets > 0:
                message = f"Completed port reset scan: {successful_resets} ports reset ({verified_resets} verified, {unverified_resets} still unreachable) out of {total_down} unreachable locations"
            else:
                message = f"Completed port reset scan: {successful_resets} ports reset but none verified as reachable out of {total_down} unreachable locations"
        else:
            message = f"No ports needed resetting: {total_checked} locations checked"
        
        # Determine overall success
        overall_success = len(reset_failed) == 0 or successful_resets > 0
        status_code = 200 if overall_success else 500
        
        response = ResponseModel(
            success=overall_success,
            message=message,
            data={
                'total_checked': total_checked,
                'down_locations': down_locations,
                'reset_attempted': reset_attempted,
                'reset_successful': reset_successful,
                'reset_failed': reset_failed,
                'not_down': not_down
            }
        )
        
        return jsonify(response.to_dict()), status_code
        
    except Exception as e:
        response = ResponseModel(
            success=False,
            message=f"Unexpected error during reset_down_port_only operation: {str(e)}",
            data={'error_type': 'UNEXPECTED_ERROR'}
        )
        return jsonify(response.to_dict()), 500


@api_bp.route('/network/reset_down_port_only_sse', methods=['POST'])
def reset_down_port_only_sse():
    """
    SSE version of reset_down_port_only - Stream reset operations in real-time
    
    Args:
        request.json (dict, optional): JSON payload containing:
            - timeout (int, optional): Connection timeout in seconds for each ping and reset
                Defaults to 10
    
    Returns:
        Server-Sent Events stream with progress updates
    """
    # Capture request data before entering generator to avoid request context issues
    data = request.json if request.json else {}
    timeout = data.get('timeout', 10)  # Default timeout of 10 seconds
    
    def generate_reset_events():
        try:
            # Load environment variables
            load_dotenv()
            
            # Validate timeout parameter
            try:
                validated_timeout = int(timeout)
                if validated_timeout <= 0 or validated_timeout > 300:
                    raise ValueError("Timeout must be between 1 and 300 seconds")
            except (ValueError, TypeError):
                yield f"data: {json.dumps({'type': 'error', 'message': 'Invalid timeout value. Must be an integer between 1 and 300 seconds', 'timestamp': time.time()})}\n\n"
                return
            
            # Check if SSH credentials are available
            ssh_username = os.getenv('SSH_USERNAME')
            ssh_password = os.getenv('SSH_PASSWORD')
            
            if not ssh_username or not ssh_password:
                yield f"data: {json.dumps({'type': 'error', 'message': 'SSH credentials not configured. Please set SSH_USERNAME and SSH_PASSWORD in .env file', 'timestamp': time.time()})}\n\n"
                return
            
            # Get all locations from data.xlsx using the shared function
            success, result = get_locations_from_data_file()
            
            if not success:
                yield f"data: {json.dumps({'type': 'error', 'message': result, 'timestamp': time.time()})}\n\n"
                return
            
            # Extract locations and dataframe from the result
            locations = result['locations']
            df_clean = result['df_clean']
            
            yield f"data: {json.dumps({'type': 'start', 'message': f'Starting Reset Down Port Only operation. Checking {len(locations)} locations...', 'timestamp': time.time()})}\n\n"
            
            total_checked = len(locations)
            down_locations = []
            reset_successful = []
            reset_failed = []
            not_down = []
            
            # Check each location
            for i, location in enumerate(locations):
                try:
                    # Get the IP for this location
                    location_data = df_clean[df_clean['Location'] == location]
                    if location_data.empty:
                        continue
                    
                    ip_address = location_data['IP'].iloc[0]
                    ip_address = str(ip_address).strip()
                    
                    # Send ping attempt status
                    yield f"data: {json.dumps({'type': 'progress', 'message': f'Pinging {location} ({ip_address}) with 3 packets, 3s timeout...', 'location': location, 'timestamp': time.time()})}\n\n"
                    
                    # Ping the IP address with more robust parameters
                    # Force English output to avoid localization issues
                    env = os.environ.copy()
                    env['LC_ALL'] = 'C'
                    ping_result = subprocess.run(
                        ['ping', ip_address], 
                        capture_output=True, 
                        text=True,
                        timeout=validated_timeout + 5,
                        env=env
                    )
                    
                    # More robust ping result validation using universal indicators
                    ping_success = False
                    if ping_result.returncode == 0:
                        # Check for universal success indicators that work across languages
                        ping_output = ping_result.stdout.lower()
                        # Look for TTL (Time To Live) which indicates successful ping response
                        if 'ttl=' in ping_output:
                            ping_success = True
                        # Alternative: look for timing indicators (ms) which also indicate success
                        elif 'ms' in ping_output and ('time=' in ping_output or '時間=' in ping_output):
                            ping_success = True
                    
                    if ping_success:
                        # IP is reachable - no action needed
                        not_down.append({'location': location, 'ip': ip_address})
                        yield f"data: {json.dumps({'type': 'ping_result', 'success': True, 'message': f'{location} is reachable - no reset needed', 'location': location, 'timestamp': time.time()})}\n\n"
                    else:
                        # IP is not reachable - attempt port reset
                        down_locations.append({'location': location, 'ip': ip_address})
                        
                        # Log ping failure details for debugging
                        ping_debug_info = f"Ping failed - Return code: {ping_result.returncode}, Output: {ping_result.stdout[:200]}..."
                        yield f"data: {json.dumps({'type': 'ping_debug', 'message': ping_debug_info, 'location': location, 'timestamp': time.time()})}\n\n"
                        
                        yield f"data: {json.dumps({'type': 'reset_attempt', 'message': f'{location} is unreachable - attempting port reset', 'location': location, 'timestamp': time.time()})}\n\n"
                        
                        try:
                            # Get SSH configuration
                            config = retrieve_ssh_info_from_config(location)
                            
                            if not config:
                                reset_failed.append({'location': location, 'error': 'SSH configuration not found'})
                                yield f"data: {json.dumps({'type': 'reset_error', 'message': 'SSH configuration not found', 'location': location, 'timestamp': time.time()})}\n\n"
                            else:
                                # Create SSH configuration
                                ssh_config = {
                                    'hostname': config['hostname'],
                                    'username': ssh_username,
                                    'password': ssh_password,
                                    'port': int(config['port'])
                                }
                                
                                target_port = config['target_port']
                                
                                # Execute port reset
                                success = reset_port_poe(ssh_config, target_port, validated_timeout)
                                
                                if success:
                                    reset_successful.append({'location': location, 'ip': ip_address, 'target_port': target_port, 'verified': False})
                                    yield f"data: {json.dumps({'type': 'reset_success', 'message': f'Successfully reset port {target_port} - verification will be performed after all resets', 'location': location, 'timestamp': time.time()})}\n\n"
                                else:
                                    reset_failed.append({'location': location, 'error': 'Port reset failed'})
                                    yield f"data: {json.dumps({'type': 'reset_error', 'message': 'Port reset failed', 'location': location, 'timestamp': time.time()})}\n\n"
                        
                        except Exception as e:
                            reset_failed.append({'location': location, 'error': f'Exception during reset: {str(e)}'})
                            yield f"data: {json.dumps({'type': 'reset_error', 'message': f'Exception during reset: {str(e)}', 'location': location, 'timestamp': time.time()})}\n\n"
                            
                except subprocess.TimeoutExpired:
                    down_locations.append({'location': location, 'ip': ip_address})
                    yield f"data: {json.dumps({'type': 'ping_result', 'success': False, 'message': f'Ping timeout for {location} - treating as unreachable', 'location': location, 'timestamp': time.time()})}\n\n"
                    
                except Exception as e:
                    reset_failed.append({'location': location, 'error': f'Error during processing: {str(e)}'})
                    yield f"data: {json.dumps({'type': 'reset_error', 'message': f'Processing error: {str(e)}', 'location': location, 'timestamp': time.time()})}\n\n"
            
            

            # Perform batch verification after all resets are completed
            successful_resets = len(reset_successful)
            total_down = len(down_locations)
            
            if successful_resets > 0:
                # Wait for devices to fully resume after resets
                batch_verification_delay = int(os.getenv('BATCH_VERIFICATION_DELAY_SECONDS', '30'))
                yield f"data: {json.dumps({'type': 'batch_verification_start', 'message': f'All {successful_resets} port resets completed. Waiting {batch_verification_delay} seconds for devices to resume before verification...', 'timestamp': time.time()})}\n\n"
                time.sleep(batch_verification_delay)
                
                # Now verify all reset ports
                verified_count = 0
                for reset_entry in reset_successful:
                    location = reset_entry['location']
                    ip_address = reset_entry['ip']
                    target_port = reset_entry['target_port']
                    
                    yield f"data: {json.dumps({'type': 'verification_progress', 'message': f'Verifying {location} ({ip_address})...', 'location': location, 'timestamp': time.time()})}\n\n"
                    
                    verification_success = verify_port_connectivity(ip_address, validated_timeout)
                    if verification_success:
                        reset_entry['verified'] = True
                        verified_count += 1
                        log.log_to_file(f"Reset Down Port Only: Ping successful for {location} ({ip_address})", 'INFO')
                        yield f"data: {json.dumps({'type': 'verification_success', 'message': f'{location} is now reachable after reset', 'location': location, 'timestamp': time.time()})}\n\n"
                    else:
                        reset_entry['verified'] = False
                        log.log_to_file(f"Reset Down Port Only: Ping failed for {location} ({ip_address})", 'ERROR')
                        yield f"data: {json.dumps({'type': 'verification_failed', 'message': f'{location} is still unreachable after reset', 'location': location, 'timestamp': time.time()})}\n\n"
                
                # Send final completion event with verification results
                unverified_count = successful_resets - verified_count
                if verified_count == successful_resets:
                    completion_message = f"Reset Down Port Only completed: {successful_resets} ports reset and all verified successfully out of {total_down} unreachable locations"
                elif verified_count > 0:
                    completion_message = f"Reset Down Port Only completed: {successful_resets} ports reset ({verified_count} verified, {unverified_count} still unreachable) out of {total_down} unreachable locations"
                else:
                    completion_message = f"Reset Down Port Only completed: {successful_resets} ports reset but none verified as reachable out of {total_down} unreachable locations"
                yield f"data: {json.dumps({'type': 'complete', 'success': True, 'message': completion_message, 'timestamp': time.time()})}\n\n"
            else:
                completion_message = f"Reset Down Port Only completed: No ports needed resetting ({total_checked} locations checked)"
                yield f"data: {json.dumps({'type': 'complete', 'success': True, 'message': completion_message, 'timestamp': time.time()})}\n\n"
        
        except Exception as e:
            yield f"data: {json.dumps({'type': 'error', 'message': f'Unexpected error during reset down port only SSE operation: {str(e)}', 'timestamp': time.time()})}\n\n"
    
    return Response(
        generate_reset_events(),
        mimetype='text/event-stream',
        headers={
            'Cache-Control': 'no-cache',
            'Connection': 'keep-alive',
            'Access-Control-Allow-Origin': '*',
            'Access-Control-Allow-Headers': 'Cache-Control'
        }
    )


@api_bp.route('/network/reset_all_locations_sse', methods=['POST'])
def reset_all_locations_sse():
    """
    SSE version of reset_all_locations - Stream reset operations in real-time
    
    Args:
        request.json (dict, optional): JSON payload containing:
            - timeout (int, optional): Connection timeout in seconds for each reset
                Defaults to 10
    
    Returns:
        Server-Sent Events stream with progress updates
    """
    # Capture request data before entering generator function
    data = request.json if request.json else {}
    timeout = data.get('timeout', 10)  # Default timeout of 10 seconds
    
    def generate_reset_events():
        try:
            # Load environment variables
            load_dotenv()
            
            # Validate timeout parameter
            try:
                timeout_val = int(timeout)
                if timeout_val <= 0 or timeout_val > 300:
                    raise ValueError("Timeout must be between 1 and 300 seconds")
            except (ValueError, TypeError):
                yield f"data: {json.dumps({'type': 'error', 'message': 'Invalid timeout value. Must be an integer between 1 and 300 seconds', 'timestamp': time.time()})}\n\n"
                return
            
            # Check if SSH credentials are available
            ssh_username = os.getenv('SSH_USERNAME')
            ssh_password = os.getenv('SSH_PASSWORD')
            
            if not ssh_username or not ssh_password:
                yield f"data: {json.dumps({'type': 'error', 'message': 'SSH credentials not configured. Please set SSH_USERNAME and SSH_PASSWORD in .env file', 'timestamp': time.time()})}\n\n"
                return
            
            # Get all locations from data.xlsx using shared function
            success, result = get_locations_from_data_file()
            
            if not success:
                yield f"data: {json.dumps({'type': 'error', 'message': result, 'timestamp': time.time()})}\n\n"
                return
            
            # Extract locations from the result
            locations = result['locations']
            
            yield f"data: {json.dumps({'type': 'start', 'message': f'Starting Reset All operation. Processing {len(locations)} locations...', 'timestamp': time.time()})}\n\n"
            
            successful_locations = []
            failed_locations = []
            
            # Reset each location
            for i, location in enumerate(locations):
                try:
                    yield f"data: {json.dumps({'type': 'reset_attempt', 'message': f'Reset attempt {i+1}/{len(locations)}', 'location': location, 'timestamp': time.time()})}\n\n"
                    
                    # Get SSH configuration
                    config = retrieve_ssh_info_from_config(location)
                    
                    if not config:
                        failed_locations.append({'location': location, 'error': 'SSH configuration not found'})
                        yield f"data: {json.dumps({'type': 'reset_error', 'message': 'SSH configuration not found', 'location': location, 'timestamp': time.time()})}\n\n"
                        continue
                    
                    # Create SSH configuration
                    ssh_config = {
                        'hostname': config['hostname'],
                        'username': ssh_username,
                        'password': ssh_password,
                        'port': int(config['port'])
                    }
                    
                    target_port = config['target_port']
                    
                    # Execute port reset
                    success = reset_port_poe(ssh_config, target_port, timeout_val)
                    
                    if success:
                        successful_locations.append(location)
                        yield f"data: {json.dumps({'type': 'reset_success', 'message': f'Successfully reset port {target_port}', 'location': location, 'timestamp': time.time()})}\n\n"
                    else:
                        failed_locations.append({'location': location, 'error': 'Port reset failed'})
                        yield f"data: {json.dumps({'type': 'reset_error', 'message': 'Port reset failed', 'location': location, 'timestamp': time.time()})}\n\n"
                    
                    # Small delay between resets
                    if i < len(locations) - 1:
                        yield f"data: {json.dumps({'type': 'progress', 'message': 'Waiting before next reset...', 'timestamp': time.time()})}\n\n"
                        time.sleep(2)
                        
                except Exception as e:
                    failed_locations.append({'location': location, 'error': str(e)})
                    yield f"data: {json.dumps({'type': 'reset_error', 'message': f'Exception: {str(e)}', 'location': location, 'timestamp': time.time()})}\n\n"
            
            # Build completion message
            success_count = len(successful_locations)
            failed_count = len(failed_locations)
            total_processed = len(locations)
            
            if success_count == total_processed:
                overall_success = True
                summary_message = f"Successfully reset {success_count} out of {total_processed} locations"
            else:
                overall_success = success_count > 0
                summary_message = f"Reset operation completed: {success_count} successful, {failed_count} failed out of {total_processed} locations"
            
            yield f"data: {json.dumps({'type': 'complete', 'success': overall_success, 'message': summary_message, 'timestamp': time.time()})}\n\n"
            
        except Exception as e:
            yield f"data: {json.dumps({'type': 'error', 'message': f'Unexpected error during reset all SSE operation: {str(e)}', 'timestamp': time.time()})}\n\n"
    
    return Response(
        generate_reset_events(),
        mimetype='text/event-stream',
        headers={
            'Cache-Control': 'no-cache',
            'Connection': 'keep-alive',
            'Access-Control-Allow-Origin': '*',
            'Access-Control-Allow-Headers': 'Cache-Control'
        }
    )


@api_bp.route('/network/ping_all_status', methods=['GET'])
def ping_all_status():
    """
    Ping all IP addresses from the data.xlsx file and return their status
    
    Returns:
        ResponseModel: JSON response containing:
            - success (bool): Whether the operation was successful
            - message (str): Human-readable message describing the result
            - data (dict): Details of the ping results including:
                - locations: List of location names
                - ips: List of IP addresses
                - statuses: List of ping statuses (True/False for each IP)
                - ping_results: Detailed ping results for each IP
    """
    try:
        # Get all locations from data.xlsx using shared function
        success, result = get_locations_from_data_file()
        
        if not success:
            response = ResponseModel(
                success=False,
                message=result,
                data=None
            )
            return jsonify(response.to_dict()), 500
        
        # Extract locations and dataframe from the result
        df_clean = result['df_clean']
        
        # Process all locations
        locations = []
        ips = []
        statuses = []
        ping_results = []
        
        # Check each location
        for _, row in df_clean.iterrows():
            location = row['Location']
            ip_address = str(row['IP']).strip()
            
            locations.append(location)
            ips.append(ip_address)
            
            # Ping the IP address
            try:
                # Force English output to avoid localization issues
                env = os.environ.copy()
                env['LC_ALL'] = 'C'
                
                ping_result = subprocess.run(
                    ['ping', ip_address], 
                    capture_output=True, 
                    text=True,
                    timeout=5,
                    env=env
                )
                
                # Check for universal success indicators that work across languages
                ping_success = False
                if ping_result.returncode == 0:
                    ping_output = ping_result.stdout.lower()
                    # Look for TTL (Time To Live) which indicates successful ping response
                    if 'ttl=' in ping_output:
                        ping_success = True
                    # Alternative: look for timing indicators (ms) which also indicate success
                    elif 'ms' in ping_output and ('time=' in ping_output or '時間=' in ping_output):
                        ping_success = True
                
                statuses.append(ping_success)
                ping_results.append({
                    'success': ping_success,
                    'return_code': ping_result.returncode,
                    'output': ping_result.stdout[:200] if ping_result.stdout else '',
                    'error': ping_result.stderr[:200] if ping_result.stderr else ''
                })
                
            except subprocess.TimeoutExpired:
                statuses.append(False)
                ping_results.append({
                    'success': False,
                    'return_code': -1,
                    'output': '',
                    'error': 'Ping timeout'
                })
            except Exception as e:
                statuses.append(False)
                ping_results.append({
                    'success': False,
                    'return_code': -1,
                    'output': '',
                    'error': f'Ping error: {str(e)}'
                })
        
        # Count successful pings
        successful_pings = sum(statuses)
        total_pings = len(statuses)
        
        response = ResponseModel(
            success=True,
            message=f"Ping status check completed: {successful_pings}/{total_pings} locations reachable",
            data={
                'locations': locations,
                'ips': ips,
                'statuses': statuses,
                'ping_results': ping_results,
                'summary': {
                    'total': total_pings,
                    'successful': successful_pings,
                    'failed': total_pings - successful_pings
                }
            }
        )
        
        return jsonify(response.to_dict())
        
    except Exception as e:
        response = ResponseModel(
            success=False,
            message=f"Unexpected error during ping status check: {str(e)}",
            data={'error_type': 'UNEXPECTED_ERROR'}
        )
        return jsonify(response.to_dict()), 500

@api_bp.route('/network/ping_single_status', methods=['POST'])
def ping_single_status():
    """
    Ping a single IP address and return its status
    
    Args:
        request.json (dict): JSON payload containing:
            - ip (str): IP address to ping
            - location (str, optional): Location name for logging
    
    Returns:
        ResponseModel: JSON response containing:
            - success (bool): Whether the operation was successful
            - message (str): Human-readable message describing the result
            - data (dict): Details of the ping result including:
                - ip: The IP address that was pinged
                - location: The location name (if provided)
                - status: Boolean indicating if ping was successful
                - ping_result: Detailed ping result
    """
    try:
        data = request.json
        if not data or 'ip' not in data:
            response = ResponseModel(
                success=False,
                message="IP address is required",
                data=None
            )
            return jsonify(response.to_dict()), 400
        
        ip_address = data['ip']
        location = data.get('location', 'Unknown')
        
        # Ping the IP address
        try:
            # Force English output to avoid localization issues
            env = os.environ.copy()
            env['LC_ALL'] = 'C'
            
            ping_result = subprocess.run(
                ['ping', ip_address], 
                capture_output=True, 
                text=True,
                timeout=5,
                env=env
            )
            
            # Check for universal success indicators that work across languages
            ping_success = False
            if ping_result.returncode == 0:
                ping_output = ping_result.stdout.lower()
                # Look for TTL (Time To Live) which indicates successful ping response
                if 'ttl=' in ping_output:
                    ping_success = True
                # Alternative: look for timing indicators (ms) which also indicate success
                elif 'ms' in ping_output and ('time=' in ping_output or '時間=' in ping_output):
                    ping_success = True
            
            response = ResponseModel(
                success=True,
                message=f"Ping completed for {location} ({ip_address})",
                data={
                    'ip': ip_address,
                    'location': location,
                    'status': ping_success,
                    'ping_result': {
                        'success': ping_success,
                        'return_code': ping_result.returncode,
                        'output': ping_result.stdout[:200] if ping_result.stdout else '',
                        'error': ping_result.stderr[:200] if ping_result.stderr else ''
                    }
                }
            )
            
            return jsonify(response.to_dict())
            
        except subprocess.TimeoutExpired:
            response = ResponseModel(
                success=True,
                message=f"Ping timeout for {location} ({ip_address})",
                data={
                    'ip': ip_address,
                    'location': location,
                    'status': False,
                    'ping_result': {
                        'success': False,
                        'return_code': -1,
                        'output': '',
                        'error': 'Ping timeout'
                    }
                }
            )
            return jsonify(response.to_dict())
            
        except Exception as e:
            response = ResponseModel(
                success=True,
                message=f"Ping error for {location} ({ip_address}): {str(e)}",
                data={
                    'ip': ip_address,
                    'location': location,
                    'status': False,
                    'ping_result': {
                        'success': False,
                        'return_code': -1,
                        'output': '',
                        'error': f'Ping error: {str(e)}'
                    }
                }
            )
            return jsonify(response.to_dict())
        
    except Exception as e:
        response = ResponseModel(
            success=False,
            message=f"Unexpected error during single ping status check: {str(e)}",
            data={'error_type': 'UNEXPECTED_ERROR'}
        )
        return jsonify(response.to_dict()), 500

@api_bp.route('/config/edit', methods=['POST'])
def edit_config():
    """
    Edit the Port assignment sheet in the config file
    """
    try:
        data = request.json
        
        if not data or 'data' not in data:
            return jsonify({'error': 'No data provided'}), 400
        
        config_data = data['data']
        
        # Validate the data structure
        if not isinstance(config_data, list):
            return jsonify({'error': 'Data must be a list of records'}), 400
        
        # Validate each record is a dictionary (no required field checks)
        for i, record in enumerate(config_data):
            if not isinstance(record, dict):
                return jsonify({'error': f'Record {i} must be a dictionary'}), 400
        
        # Get the project root directory
        project_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        excel_path = os.path.join(project_root, 'config_file', 'data.xlsx')
        
        # Read the existing Excel file
        try:
            # Load the existing workbook
            from openpyxl import load_workbook
            workbook = load_workbook(excel_path)
            
            # Check if 'Port assignment' sheet exists, create if not
            if 'Port assignment' not in workbook.sheetnames:
                workbook.create_sheet('Port assignment')
            
            worksheet = workbook['Port assignment']
            
            # Clear existing data (except header row)
            if worksheet.max_row > 1:
                worksheet.delete_rows(2, worksheet.max_row)
            
            # Set headers if sheet is empty
            if worksheet.max_row == 0:
                headers = ['Unnamed: 0', 'Switch Port', 'Location', 'SSH IP']
                for col, header in enumerate(headers, 1):
                    worksheet.cell(row=1, column=col, value=header)
            
            # Write the new data
            for row_idx, record in enumerate(config_data, 2):  # Start from row 2 (after header)
                worksheet.cell(row=row_idx, column=1, value=record.get('Unnamed: 0', ''))
                worksheet.cell(row=row_idx, column=2, value=record.get('Switch Port', ''))
                worksheet.cell(row=row_idx, column=3, value=record.get('Location', ''))
                worksheet.cell(row=row_idx, column=4, value=record.get('SSH IP', ''))
            
            # Save the workbook
            workbook.save(excel_path)
            workbook.close()
            
            log.log_to_file(f"Port assignment config updated successfully with {len(config_data)} records")
            return jsonify({'message': 'Port assignment config edited successfully', 'records_updated': len(config_data)}), 200
            
        except Exception as e:
            log.log_to_file(f"Error updating Excel file: {str(e)}", 'ERROR')
            return jsonify({'error': f'Failed to update config file: {str(e)}'}), 500
            
    except Exception as e:
        log.log_to_file(f"Error in edit_config: {str(e)}", 'ERROR')
        return jsonify({'error': f'Internal server error: {str(e)}'}), 500


@api_bp.route('/config/edit_hardware', methods=['POST'])
def edit_hardware_config():
    """
    Edit the Hardware list sheet in the config file
    """
    try:
        data = request.json
        
        if not data or 'data' not in data:
            return jsonify({'error': 'No data provided'}), 400
        
        hardware_data = data['data']
        
        # Validate the data structure
        if not isinstance(hardware_data, list):
            return jsonify({'error': 'Data must be a list of records'}), 400
        
        # Validate each record is a dictionary (no required field checks)
        for i, record in enumerate(hardware_data):
            if not isinstance(record, dict):
                return jsonify({'error': f'Record {i} must be a dictionary'}), 400
        
        # Get the project root directory
        project_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        excel_path = os.path.join(project_root, 'config_file', 'data.xlsx')
        
        # Read the existing Excel file
        try:
            # Load the existing workbook
            from openpyxl import load_workbook
            workbook = load_workbook(excel_path)
            
            # Check if 'Hardware list' sheet exists, create if not
            if 'Hardware list' not in workbook.sheetnames:
                workbook.create_sheet('Hardware list')
            
            worksheet = workbook['Hardware list']
            
            # Clear existing data (except header row)
            if worksheet.max_row > 1:
                worksheet.delete_rows(2, worksheet.max_row)
            
            # Use explicit column order to maintain consistency
            headers = ['Location', 'IP']
            
            # Set headers if sheet is empty
            if worksheet.max_row == 0:
                for col, header in enumerate(headers, 1):
                    worksheet.cell(row=1, column=col, value=header)
            
            # Write the new data
            for row_idx, record in enumerate(hardware_data, 2):  # Start from row 2 (after header)
                for col_idx, header in enumerate(headers, 1):
                    worksheet.cell(row=row_idx, column=col_idx, value=record.get(header, ''))
            
            # Save the workbook
            workbook.save(excel_path)
            workbook.close()
            
            log.log_to_file(f"Hardware list config updated successfully with {len(hardware_data)} records")
            return jsonify({'message': 'Hardware list config edited successfully', 'records_updated': len(hardware_data)}), 200
            
        except Exception as e:
            log.log_to_file(f"Error updating Excel file: {str(e)}", 'ERROR')
            return jsonify({'error': f'Failed to update config file: {str(e)}'}), 500
            
    except Exception as e:
        log.log_to_file(f"Error in edit_hardware_config: {str(e)}", 'ERROR')
        return jsonify({'error': f'Internal server error: {str(e)}'}), 500





